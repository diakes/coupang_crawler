import requests
import re
from bs4 import BeautifulSoup
from setuptools.package_index import HREF
from selenium import webdriver
from openpyxl import load_workbook # 파일 불러오기
from openpyxl import Workbook
#https://www.coupang.com/vp/products/315169790?itemId=1001641511&vendorItemId=72519465705&q=%ED%8C%A8%EB%94%A9&itemsCount=36&searchId=7c075a41d29742cbab8ed334163b232d&rank=9&isAddedCart=
f =open("my88.lxml", "r", encoding="utf8")

da=f.read()
item = BeautifulSoup(da)


name = item.find("h2", attrs={"class":"prod-buy-header__title"}).get_text() # 제품명
print(f"제품명 : {name}")

price = item.find("span", attrs={"class":"total-price"}).get_text().strip() # 가격
print(f"가격 : {price}")

rate = item.find("span", attrs={"class":"rating-star-container"}) # 평점
if rate:#평점
    rate = rate.span["style"][-6:-2]
else:
    rate = "없음"
    
rate_cnt = item.find("span", attrs={"class":"count"}) # 평점 수 
if rate_cnt:
    rate_cnt = rate_cnt.get_text()[0:3] # 예 : (26)
else:
   rate_cnt = "평점 수 없음"
print(f"평점 : {rate}점 ({rate_cnt}개)")

ing = item.find("img", attrs={"class":"prod-image__detail"})["src"] # 대표이미지
print(f"이미지 : {ing}")

Kategorie= item.find_all("a", attrs={"class":"breadcrumb-link"})# 카테고리
i=0
for Kat in Kategorie:
    Kategorie[i] = Kat.get_text().strip()
    i+=1
print(f"카테고리: {Kategorie}")

Detailed = item.find("table", attrs={"class":"prod-delivery-return-policy-table essential-info-table"})# 상세정보
#print(f"상세정보 : {Detailed}")

code = item.find("meta", attrs={"property":"og:url"})["content"].strip()# 상품코드
code2=code.replace("https://www.coupang.com/vp/products/", "").strip()
print(f"상품코드 : {code2}")

Producers = item.find("table", attrs={"class":"prod-delivery-return-policy-table essential-info-table"})# 판매자명**
Producers = Producers.find("th", text="생산자(수입자)").find_next("td").get_text()
print(f"원산지 : {Producers}")

Country = item.find("table", attrs={"class":"prod-delivery-return-policy-table essential-info-table"})# 판매자명**
Country = Country.find("th", text="원산지").find_next("td").get_text()
print(f"원산지 : {Country}")

Shipping = item.find("th", text="배송비").find_next("td").get_text().strip()
#Shipping = item.find("div", attrs={"class":"prod-shipping-fee-message"}).get_text().strip()# 배송 
p = re.compile("무료배송")
Shipping =p.search(Shipping).group()
print(f"배송 : {Shipping}")

print("-"*100) # 줄긋기
wb = load_workbook("shoppling_prod_bluk_reg_sample.xlsx") # sample.xlsx 파일에서 wb 을 불러옴
ws = wb.active # 활성화된 Sheet
ws["A4"] = name #상품명
ws["B4"] = code2 #자사상품 코드
ws["K4"] = "B" #판매상태
ws["L4"] =  Producers #제조사명
ws["s4"] = Country #원산지
ws["Y4"] = "A"#과세정보

if Shipping =="무료배송":
    Shipping ="A"
elif Shipping =="착불":
    Shipping ="B"
elif Shipping =="선결제":
    Shipping ="C"
elif Shipping =="조건부배송":
    Shipping ="D"
else:
    Shipping ="z"

ws["Z4"] = Shipping#배송비구분
ws["AX4"] = Shipping#샵플링 소지바가
ws["AY4"] = price# 샤플링 판매가
ws["AZ4"] = price#샵프링 매입가
ws["BG4"] = ing#대표 이미지
#ws["CN4"] = Detailed[0:51] #상세 설명
c=""
for a in Kategorie:
    c=c+a+">"
print(c)
ws["dx4"] = c[0:-1]#표준카테고리명

wb.save("shoppling_prod_bluk_reg_sample2(005).xlsx")
wb.close()


    